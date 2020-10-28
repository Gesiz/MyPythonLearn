import openpyxl

wb = openpyxl.load_workbook('example.xlsx')  # 打开工作簿并生成类型对象 example.xlsx 为已存在工作簿

# Getting sheets from the workbook
print(wb.sheetnames)  # 打印表单名称

for sheet in wb:  # 遍历表单
    print(sheet.title)

# mySheet = wb.create_sheet('mySheet')  # 在后尾插入
# ws2 = wb.create_sheet("Mysheet", 0)
print(wb.sheetnames)

# sheet1 = wb.get_sheet_by_name('Sheet1')
# sheet1 = wb['sheet1']
# ws = wb.active  # 激活当前表
ws = wb['Sheet1']  # 激活worksheet
print(ws)

print(ws['A1'])
print(ws['A1'].value)

c = ws['B1']
print(f"Row {c.row}, Column {c.column} is {c.value}")